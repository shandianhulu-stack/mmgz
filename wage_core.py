from __future__ import annotations

import math
import os
import sys
from dataclasses import dataclass
from decimal import Decimal, ROUND_FLOOR, ROUND_HALF_UP
from typing import Dict, Iterable, List, Optional, Tuple

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
VENDOR_DIR = os.path.join(BASE_DIR, "vendor")
if os.path.isdir(VENDOR_DIR) and VENDOR_DIR not in sys.path:
    sys.path.insert(0, VENDOR_DIR)

VERSION = "Final_V1.1"

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

UNIT_AMOUNT = Decimal("700")
RATING_STEP_AMOUNT = Decimal("0.07")  # 700 * 0.01%

SHEET1_HEADERS = [
    "姓名", "电话", "金额", "预算", "差额（预算-预付 ）", "已完成场次",
    "本次结算场次", "本次可结算最大总金额", "本次可计算金额是否大于或等于实际结算金额",
    "五级评价后实付金额", "差值（五级评价后实付金额-预付）",
]


@dataclass
class PersonInput:
    name: str
    phone: str
    amount: Decimal


@dataclass
class SettlementInfo:
    budget_formula_or_value: object = None
    completed_count: object = None
    settle_count: object = None
    max_total: object = None
    can_calc: object = None


def _norm_header(v) -> str:
    return str(v).replace(" ", "").replace("\n", "").strip() if v is not None else ""


def _norm_phone(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace(" ", "")


def _to_decimal(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0.00")
    return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _to_int_or_none(v) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(Decimal(str(v)))
    except Exception:
        return None


def _find_header_row(ws, required: Iterable[str], max_scan_rows: int = 30) -> Tuple[int, Dict[str, int]]:
    required = list(required)
    for row in range(1, min(ws.max_row, max_scan_rows) + 1):
        mapping = {}
        for col in range(1, ws.max_column + 1):
            h = _norm_header(ws.cell(row, col).value)
            if h:
                mapping[h] = col
        if all(k in mapping for k in required):
            return row, mapping
    raise ValueError(f"没有找到表头：{required}")


def read_mimo_score(path: str) -> List[PersonInput]:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row, mapping = _find_header_row(ws, ["姓名", "电话", "金额"])
    result: List[PersonInput] = []
    for row in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row, mapping["姓名"]).value
        phone = ws.cell(row, mapping["电话"]).value
        amount = ws.cell(row, mapping["金额"]).value
        if name is None and phone is None and amount is None:
            continue
        if not name or amount is None:
            continue
        result.append(PersonInput(str(name).strip(), _norm_phone(phone), _to_decimal(amount)))
    if not result:
        raise ValueError("米墨评分表没有读取到有效数据")
    return result


def read_settlement(path: Optional[str]) -> Dict[Tuple[str, str], SettlementInfo]:
    """读取用户上传的结算过程表/成品工资明细表。

    这个表用于回填 Sheet1 需要和原始成品一致的过程字段：
    预算、已完成场次、本次结算场次、本次可结算最大总金额、是否足额。
    """
    if not path:
        return {}
    wb_formula = load_workbook(path, data_only=False)
    wb_value = load_workbook(path, data_only=True)
    ws_f = wb_formula[wb_formula.sheetnames[0]]
    ws_v = wb_value[wb_value.sheetnames[0]]
    header_row, mapping = _find_header_row(ws_f, ["姓名", "电话"])
    ref: Dict[Tuple[str, str], SettlementInfo] = {}
    col = lambda h: mapping.get(_norm_header(h))
    for row in range(header_row + 1, ws_f.max_row + 1):
        name = ws_f.cell(row, mapping["姓名"]).value
        phone = ws_f.cell(row, mapping["电话"]).value
        if not name:
            continue
        key = (str(name).strip(), _norm_phone(phone))
        info = SettlementInfo()
        if col("预算"):
            info.budget_formula_or_value = ws_v.cell(row, col("预算")).value
        if col("已完成场次"):
            info.completed_count = ws_v.cell(row, col("已完成场次")).value
        if col("本次结算场次"):
            info.settle_count = ws_v.cell(row, col("本次结算场次")).value
        if col("本次可结算最大总金额"):
            info.max_total = ws_v.cell(row, col("本次可结算最大总金额")).value
        if col("本次可计算金额是否大于或等于实际结算金额"):
            info.can_calc = ws_v.cell(row, col("本次可计算金额是否大于或等于实际结算金额")).value
        ref[key] = info
    return ref


def floor_to_rating_step(amount: Decimal) -> Decimal:
    if amount <= 0:
        return Decimal("0.00")
    steps = (amount / RATING_STEP_AMOUNT).to_integral_value(rounding=ROUND_FLOOR)
    return (steps * RATING_STEP_AMOUNT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calc_default_scene_count(amount: Decimal) -> int:
    payout = floor_to_rating_step(amount)
    if payout <= 0:
        return 0
    return int(math.ceil(float(payout / UNIT_AMOUNT)))


def choose_best_payout(amount: Decimal, max_pay: Decimal, mode: str = "nearest") -> Decimal:
    """选择目标实付金额。保留给旧策略使用。"""
    if amount <= 0:
        return Decimal("0.00")
    amount = min(amount, max_pay)
    raw_steps = amount / RATING_STEP_AMOUNT
    floor_steps = int(raw_steps.to_integral_value(rounding=ROUND_FLOOR))
    candidates = {floor_steps, floor_steps + 1}
    best = None
    best_key = None
    for steps in candidates:
        if steps < 0:
            continue
        pay = (Decimal(steps) * RATING_STEP_AMOUNT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if pay > max_pay:
            continue
        if mode == "floor" and pay > amount:
            continue
        diff = pay - amount
        key = (abs(diff), 0 if diff == 0 else (1 if diff < 0 else 2))
        if best is None or key < best_key:
            best = pay
            best_key = key
    if best is None:
        return floor_to_rating_step(amount)
    return best.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _max_steps_per_scene() -> int:
    return int((UNIT_AMOUNT / RATING_STEP_AMOUNT).to_integral_value())  # 10000 = 100.00%


def _score_from_steps(steps: int) -> Decimal:
    max_steps_per_scene = _max_steps_per_scene()
    return (Decimal(max(0, min(max_steps_per_scene, int(steps)))) / Decimal(max_steps_per_scene)).quantize(Decimal("0.0001"), rounding=ROUND_FLOOR)


def _target_steps_for_amount(amount: Decimal, scene_count: int, allow_overpay: bool = False) -> int:
    """把目标金额转换成评分步数。1步=0.01%评分=0.07元。"""
    max_total_steps = _max_steps_per_scene() * max(int(scene_count or 0), 0)
    if amount <= 0 or max_total_steps <= 0:
        return 0
    raw_steps = amount / RATING_STEP_AMOUNT
    if allow_overpay:
        steps = int(raw_steps.to_integral_value(rounding=ROUND_HALF_UP))
    else:
        steps = int(raw_steps.to_integral_value(rounding=ROUND_FLOOR))
    return max(0, min(steps, max_total_steps))


def _batch_balanced_target_steps(items: List[Tuple[PersonInput, int]]) -> List[int]:
    """全批次差值贴近0的目标步数。

    单人金额被0.07约束时，逐人向下取整会把每个人的几分钱误差累加成批次误差。
    人工核算经常看的不是每个人数学下取整，而是整批差值尽量接近0。
    这里先全部向下取整，再挑一部分人+1个评分步长（+0.07元），把整批总差值压到最小。
    """
    floors: List[int] = []
    candidates = []
    floor_total_diff = Decimal("0.00")
    for idx, (p, count) in enumerate(items):
        max_total_steps = _max_steps_per_scene() * max(int(count or 0), 0)
        raw_steps = p.amount / RATING_STEP_AMOUNT if p.amount > 0 else Decimal("0")
        floor_steps = int(raw_steps.to_integral_value(rounding=ROUND_FLOOR))
        floor_steps = max(0, min(floor_steps, max_total_steps))
        floors.append(floor_steps)
        floor_pay = (Decimal(floor_steps) * RATING_STEP_AMOUNT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        floor_total_diff += floor_pay - p.amount
        ceil_steps = floor_steps + 1
        if ceil_steps <= max_total_steps:
            ceil_pay = (Decimal(ceil_steps) * RATING_STEP_AMOUNT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if ceil_pay != floor_pay:
                # 优先给“微超最少”的人+0.07，避免某个人看起来多发太多。
                candidates.append((ceil_pay - p.amount, abs(floor_pay - p.amount), idx))
    if floor_total_diff >= 0 or not candidates:
        return floors

    need = int(((-floor_total_diff) / RATING_STEP_AMOUNT).to_integral_value(rounding=ROUND_HALF_UP))
    need = max(0, min(need, len(candidates)))

    # 比较 need-1 / need / need+1，选整批总差值绝对值最小的个数。
    best_m = need
    best_key = None
    for m in {max(0, need - 1), need, min(len(candidates), need + 1)}:
        total_diff = floor_total_diff + RATING_STEP_AMOUNT * Decimal(m)
        key = (abs(total_diff), 0 if total_diff <= 0 else 1)  # 同等误差优先不超发整批
        if best_key is None or key < best_key:
            best_key = key
            best_m = m

    chosen = list(floors)
    candidates.sort(key=lambda x: (x[0], x[1]))
    for _, _, idx in candidates[:best_m]:
        chosen[idx] += 1
    return chosen


def _choose_anchor_scores(adjust_total_steps: int, k: int, min_score_steps: int) -> List[int]:
    """给k个调整行生成评分步数。

    目标是像人工：前面用常见整档（80/85/90/70/75等），最后一行精确补差。
    V0.7的问题是过分追求每行都高，导致动了太多行；V0.9改为先用最少调整行。
    """
    max_step = _max_steps_per_scene()
    if k <= 1:
        return [max(0, min(max_step, adjust_total_steps))]

    # 频次来自 2026.6.29 人工样表：80、70、85、90最常见，其次75/84/60/65/93/95。
    common = [8000, 7000, 8500, 9000, 7500, 8400, 6000, 6500, 9300, 9500, 8700, 8200, 7800, 7700, 6700, 5500, 5000]
    best = None
    best_key = None

    def dfs(prefix: List[int], depth: int):
        nonlocal best, best_key
        if depth == k - 1:
            residual = adjust_total_steps - sum(prefix)
            if residual < 0 or residual > max_step:
                return
            scores = prefix + [residual]
            if min(scores) < min_score_steps:
                return
            # 评分观感：先保证最少行已确定；这里优先常用整数档、最后一行不太低、整体不要过于平均。
            last = scores[-1]
            common_rank = sum((common.index(x) if x in common else 99) for x in prefix)
            integer_count = sum(1 for x in prefix if x % 100 == 0)
            last_band_penalty = 0 if last >= 7500 else (1 if last >= 7000 else 2)
            min_score = min(scores)
            # 不强求最高分；人工样表里70/80比一堆86.xx更常见。
            key = (
                last_band_penalty,
                -integer_count,
                common_rank,
                -min_score,
                abs(last - 8000),
            )
            if best is None or key < best_key:
                best = scores
                best_key = key
            return
        used = sum(prefix)
        slots_left = k - depth
        for score in common:
            remaining = adjust_total_steps - used - score
            if remaining < (slots_left - 1) * min_score_steps:
                continue
            if remaining > (slots_left - 1) * max_step:
                continue
            dfs(prefix + [score], depth + 1)

    dfs([], 0)
    if best is not None:
        return best

    # 兜底：均分到少量调整行，保证总额正确。
    base = adjust_total_steps // k
    extra = adjust_total_steps % k
    return [base + (1 if i < extra else 0) for i in range(k)]


def _human_anchor_score_steps(target_steps: int, scene_count: int, min_score_percent: int = 70, max_adjust_rows: int = 5) -> List[int]:
    """人工锚点法：多个100%，最后少量行调分。"""
    max_step = _max_steps_per_scene()
    scene_count = max(int(scene_count or 0), 0)
    if scene_count <= 0:
        return []
    target_steps = max(0, min(int(target_steps), scene_count * max_step))
    if target_steps >= scene_count * max_step:
        return [max_step] * scene_count

    min_score_steps = int(min_score_percent * 100)
    max_adjust_rows = max(1, min(max_adjust_rows, scene_count))

    chosen_k = max_adjust_rows
    chosen_adjust_total = target_steps - (scene_count - chosen_k) * max_step
    for k in range(1, max_adjust_rows + 1):
        adjust_total = target_steps - (scene_count - k) * max_step
        if adjust_total < 0 or adjust_total > k * max_step:
            continue
        if k == 1:
            # 一行够高就一行；低于阈值才“换2个”。
            if adjust_total >= min_score_steps or k == max_adjust_rows:
                chosen_k = k
                chosen_adjust_total = adjust_total
                break
        else:
            if adjust_total >= k * min_score_steps:
                chosen_k = k
                chosen_adjust_total = adjust_total
                break

    adjust_scores = _choose_anchor_scores(chosen_adjust_total, chosen_k, min_score_steps)
    full_rows = scene_count - chosen_k
    return [max_step] * full_rows + adjust_scores


def _legacy_balanced_score_steps(target_steps: int, scene_count: int) -> List[int]:
    max_step = _max_steps_per_scene()
    max_total_steps = max_step * scene_count
    deduction = max_total_steps - target_steps
    if deduction <= 0:
        return [max_step] * scene_count
    preferred_min = 8500
    preferred_deduct_cap = max_step - preferred_min
    if deduction <= preferred_deduct_cap * scene_count:
        base_deduct = deduction // scene_count
        extra = deduction % scene_count
        deductions = [base_deduct + (1 if i < extra else 0) for i in range(scene_count)]
        return [max_step - d for d in deductions]
    base_steps = target_steps // scene_count
    extra_steps = target_steps % scene_count
    return [base_steps + (1 if i < extra_steps else 0) for i in range(scene_count)]


def _last_fill_score_steps(target_steps: int, scene_count: int) -> List[int]:
    max_step = _max_steps_per_scene()
    scores: List[int] = []
    remaining_steps = target_steps
    for _ in range(scene_count):
        scene_steps = min(max_step, max(remaining_steps, 0))
        scores.append(scene_steps)
        remaining_steps -= scene_steps
    return scores


def split_scores_from_steps(amount: Decimal, scene_count: int, target_steps: int, strategy: str = "human_like") -> Tuple[int, Decimal, List[Decimal]]:
    scene_count = max(int(scene_count or 0), 0)
    if scene_count == 0 or amount <= 0:
        return 0, Decimal("0.00"), []
    max_total_steps = _max_steps_per_scene() * scene_count
    target_steps = max(0, min(int(target_steps), max_total_steps))
    payout = (Decimal(target_steps) * RATING_STEP_AMOUNT).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    if strategy == "last_fill":
        steps = _last_fill_score_steps(target_steps, scene_count)
    elif strategy == "zero_first_high":
        steps = _legacy_balanced_score_steps(target_steps, scene_count)
    else:
        # V0.10默认：人工锚点少量调分。是否允许微超只影响target_steps，不影响这里的拆分风格。
        steps = _human_anchor_score_steps(target_steps, scene_count, min_score_percent=70, max_adjust_rows=5)
    return scene_count, payout, [_score_from_steps(st) for st in steps]


def split_scores(amount: Decimal, scene_count: Optional[int] = None, strategy: str = "human_like") -> Tuple[int, Decimal, List[Decimal]]:
    """返回：结算场次数、最终实付金额、每场评分。

    V0.10默认策略：严格不超发 + 人工锚点少量调分。
    - 金额层：可选择给部分人+0.07元，让整批总差值接近0；
    - 明细层：多个100%，一行太低才换2/3/4/5行，不再把所有场次平均。
    """
    if scene_count is None:
        scene_count = calc_default_scene_count(amount)
    allow = strategy in ("closest_allow_overpay", "human_batch_balance")
    target_steps = _target_steps_for_amount(amount, int(scene_count or 0), allow_overpay=allow)
    return split_scores_from_steps(amount, int(scene_count or 0), target_steps, strategy=strategy)


def get_joined_info(p: PersonInput, settlement: Dict[Tuple[str, str], SettlementInfo]) -> SettlementInfo:
    return settlement.get((p.name, p.phone)) or settlement.get((p.name, "")) or SettlementInfo()


def build_preview_rows(score_path: str, settlement_path: Optional[str] = None, strategy: str = "human_like") -> Dict[str, object]:
    people = read_mimo_score(score_path)
    settlement = read_settlement(settlement_path)
    rows = []
    total_target = Decimal("0.00")
    total_payout = Decimal("0.00")
    missing_settlement = 0

    prepared = []
    for p in people:
        info = get_joined_info(p, settlement)
        uploaded_count = _to_int_or_none(info.settle_count)
        count = uploaded_count if uploaded_count is not None else calc_default_scene_count(p.amount)
        prepared.append((p, info, int(count or 0)))

    if strategy == "human_batch_balance":
        target_steps_list = _batch_balanced_target_steps([(p, count) for p, _, count in prepared])
    else:
        allow = strategy == "closest_allow_overpay"
        target_steps_list = [_target_steps_for_amount(p.amount, count, allow_overpay=allow) for p, _, count in prepared]

    overpay_count = 0
    max_abs_diff = Decimal("0.00")
    for (p, info, count), target_steps in zip(prepared, target_steps_list):
        count, payout, scores = split_scores_from_steps(p.amount, count, target_steps, strategy=strategy)
        diff = (payout - p.amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if diff > 0:
            overpay_count += 1
        if abs(diff) > max_abs_diff:
            max_abs_diff = abs(diff)

        max_total = info.max_total if info.max_total not in (None, "") else float(UNIT_AMOUNT * Decimal(count))
        try:
            can_calc = str(info.can_calc) if info.can_calc not in (None, "") else ("是" if Decimal(str(max_total)) >= p.amount else "否")
        except Exception:
            can_calc = str(info.can_calc or "")
        if not settlement_path or info.settle_count is None:
            missing_settlement += 1

        details = []
        for idx, score in enumerate(scores, start=1):
            scene_pay = (UNIT_AMOUNT * score).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            details.append({
                "index": idx,
                "name": p.name,
                "actual_amount": float(p.amount),
                "task_amount": float(UNIT_AMOUNT),
                "score": float(score),
                "score_text": f"{float(score) * 100:.2f}%",
                "pay_amount": float(scene_pay),
            })

        total_target += p.amount
        total_payout += payout
        rows.append({
            "name": p.name,
            "phone": p.phone,
            "amount": float(p.amount),
            "budget": info.budget_formula_or_value,
            "budget_diff": (float(Decimal(str(info.budget_formula_or_value)) - p.amount) if info.budget_formula_or_value not in (None, "") else None),
            "completed_count": info.completed_count,
            "settle_count": count,
            "max_total": max_total,
            "can_calc": can_calc,
            "payout": float(payout),
            "diff": float(diff),
            "abs_diff": float(abs(diff)),
            "overpay": bool(diff > 0),
            "last_score": float(scores[-1]) if scores else None,
            "last_score_text": (f"{float(scores[-1]) * 100:.2f}%" if scores else ""),
            "min_score_text": (f"{float(min(scores)) * 100:.2f}%" if scores else ""),
            "max_score_text": (f"{float(max(scores)) * 100:.2f}%" if scores else ""),
            "avg_score_text": (f"{(sum(float(x) for x in scores) / len(scores)) * 100:.2f}%" if scores else ""),
            "adjust_count": sum(1 for x in scores if x != Decimal("1.0000")),
            "strategy": strategy,
            "details": details,
        })

    total_diff = (total_payout - total_target).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return {
        "people": people,
        "settlement": settlement,
        "rows": rows,
        "count": len(rows),
        "total_target": float(total_target),
        "total_payout": float(total_payout),
        "total_diff": float(total_diff),
        "max_abs_diff": float(max_abs_diff),
        "overpay_count": overpay_count,
        "missing_settlement": missing_settlement,
    }


def _write_sheet1(wb: Workbook, preview: Dict[str, object]) -> None:
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([None] * len(SHEET1_HEADERS))
    ws.append(SHEET1_HEADERS)

    for idx, row in enumerate(preview["rows"], start=3):
        ws.cell(idx, 1, row["name"])
        ws.cell(idx, 2, row["phone"])
        ws.cell(idx, 3, row["amount"])
        if row.get("budget") is not None:
            ws.cell(idx, 4, row["budget"])
        
        if row.get("budget_diff") is not None:
            ws.cell(idx, 5, row["budget_diff"])
        if row.get("completed_count") is not None:
            ws.cell(idx, 6, row["completed_count"])
        ws.cell(idx, 7, row["settle_count"])
        ws.cell(idx, 8, row["max_total"])
        ws.cell(idx, 9, row["can_calc"])
        ws.cell(idx, 10, row["payout"])
        ws.cell(idx, 11, row["diff"])


def _write_final_sheet(wb: Workbook, preview: Dict[str, object]) -> None:
    final = wb.create_sheet("最终金额")
    final.append(["评分数据", None, None, None, None, "差值"])
    final_row = 2
    for person_idx, row in enumerate(preview["rows"]):
        # 按用户样表结构：第一组有完整表头；后续组延续样表习惯，只露出姓名/实际下发金额。
        if person_idx == 0:
            final.cell(final_row, 1, "姓名")
            final.cell(final_row, 2, "实际下发金额")
            final.cell(final_row, 3, "任务金额")
            final.cell(final_row, 4, "评分")
            final.cell(final_row, 5, "实付金额")
        else:
            final.cell(final_row, 1, "姓名")
            final.cell(final_row, 2, "实际下发金额")
        start = final_row + 1
        r = start
        for i, detail in enumerate(row["details"]):
            final.cell(r, 1, row["name"] if i == 0 else None)
            final.cell(r, 3, detail["task_amount"])
            final.cell(r, 4, detail["score"])
            final.cell(r, 5, detail["pay_amount"])
            r += 1
        final.cell(r, 1, "合计")
        final.cell(r, 2, row["amount"])
        final.cell(r, 5, row["payout"] if r > start else 0)
        final.cell(r, 6, row["diff"])
        final_row = r + 2


def _style_sheet1(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=11):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = [14, 16, 12, 12, 18, 12, 14, 18, 34, 20, 28]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:K{ws.max_row}"
    for row in range(3, ws.max_row + 1):
        ws.cell(row, 2).number_format = "@"
        for col in (3, 4, 5, 8, 10, 11):
            ws.cell(row, col).number_format = "0.00"
        for col in (6, 7):
            ws.cell(row, col).number_format = "0"


def _style_final(ws):
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=6):
        for cell in row:
            if cell.value is not None:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(1, ws.max_row + 1):
        vals = [ws.cell(row, c).value for c in range(1, 7)]
        if vals[0] in ("评分数据", "姓名"):
            for c in range(1, 7):
                ws.cell(row, c).fill = header_fill
                ws.cell(row, c).font = Font(bold=True)
    widths = [14, 16, 12, 12, 14, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for row in range(1, ws.max_row + 1):
        ws.cell(row, 2).number_format = "0.00"
        ws.cell(row, 3).number_format = "0.00"
        ws.cell(row, 4).number_format = "0.00%"
        ws.cell(row, 5).number_format = "0.00"
        ws.cell(row, 6).number_format = "0.00"


def _write_rules(wb: Workbook, settlement_path: Optional[str], strategy: str) -> None:
    rules = wb.create_sheet("规则说明")
    rules_data = [
        ["规则", "说明"],
        ["输入1", "米墨评分表：姓名、电话、金额"],
        ["输入2", "结算过程表：预算、已完成场次、本次结算场次、本次可结算最大总金额、是否足额。用于保证 Sheet1 和运营成品表格式一致。"],
        ["每场金额", "700元/场"],
        ["评分精度", "百分比显示小数点后2位，例如85.29%"],
        ["最小金额步长", "0.07元，因为700×0.01%=0.07"],
        ["实付金额", "默认按财务安全模式：严格不超发。逐人向下取整，任何人不微超；全批次差值会负数累计。"],
        ["评分分摊策略", "人工少量调分：严格不超发" if strategy == "human_like" else ("批次差值贴0：允许微超几分钱" if strategy == "human_batch_balance" else ("多行均衡高分" if strategy == "zero_first_high" else ("逐人四舍五入" if strategy == "closest_allow_overpay" else "最后一场补差（旧逻辑）")))],
        ["Final V1.1说明", "默认按用户选择B：严格不超发；同时保留多种历史/对比策略可切换。默认策略按人工习惯拆分：多个100%，一行过低才换2/3/4/5行；切换到批次贴0或逐人四舍五入时可能微超几分钱。"],
        ["导出格式", "Sheet1 按成品工资明细输出11列；最终金额输出每个人每场评分明细。"],
        ["结算过程表状态", "已上传" if settlement_path else "未上传：过程字段会按默认规则计算/留空，不能完全复刻成品表。"],
    ]
    for row in rules_data:
        rules.append(row)
    rules.column_dimensions["A"].width = 24
    rules.column_dimensions["B"].width = 100
    for cell in rules[1]:
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
        cell.font = Font(bold=True)
    for row in rules.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def generate_workbook(score_path: str, output_path: str, settlement_path: Optional[str] = None, strategy: str = "human_like") -> str:
    preview = build_preview_rows(score_path, settlement_path, strategy=strategy)
    wb = Workbook()
    _write_sheet1(wb, preview)
    _write_final_sheet(wb, preview)
    _write_rules(wb, settlement_path, strategy)
    _style_sheet1(wb["Sheet1"])
    _style_final(wb["最终金额"])
    out_dir = os.path.dirname(output_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(output_path)
    return output_path
